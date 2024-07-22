from tkinter import Toplevel, Label, Entry, Button, messagebox
import pandas as pd
import os
from openpyxl import Workbook
import numpy as np


class ManageDataValidation:
    def __init__(self, connection, database, on_import=None):
        self.connection = connection
        self.database = database
        self.tables = []
        self.errors_found = False
        self.column_errors_found = False
        self.df = None
        self.on_import = on_import

    def get_columns(self, table_name):
        cursor = self.connection.cursor()
        cursor.execute(f"SHOW COLUMNS FROM {table_name}")
        columns = [column[0] for column in cursor.fetchall()]
        cursor.close()
        return columns

    def get_column_info(self, table_name):
        cursor = self.connection.cursor()
        cursor.execute(f"SHOW COLUMNS FROM {self.database}.{table_name}")
        column_info = {column[0]: {"type": column[1], "null": column[2]} for column in cursor.fetchall()}

        cursor.execute(f"SHOW INDEX FROM {self.database}.{table_name} WHERE Non_unique = 0")
        unique_columns = [column[4] for column in cursor.fetchall()]
        for col in unique_columns:
            if col in column_info:
                column_info[col]["unique"] = True

        cursor.execute(f"""
            SELECT 
                k.COLUMN_NAME, 
                k.REFERENCED_TABLE_NAME, 
                k.REFERENCED_COLUMN_NAME 
            FROM 
                information_schema.KEY_COLUMN_USAGE k 
            WHERE 
                k.TABLE_SCHEMA = '{self.database}' 
                AND k.TABLE_NAME = '{table_name}' 
                AND k.REFERENCED_TABLE_NAME IS NOT NULL
        """)
        foreign_keys = cursor.fetchall()
        for fk in foreign_keys:
            column_info[fk[0]]['foreign_key'] = {
                'referenced_table': fk[1],
                'referenced_column': fk[2]
            }

        cursor.close()
        return column_info

    def correct_columns(self, wrong_columns, missing_columns, table_name):

        new_window = Toplevel()
        new_window.title("Correct Column Names")

        Label(new_window, text="Wrong Column").grid(row=0, column=0)
        Label(new_window, text="Correct Column").grid(row=0, column=1)

        entries = []
        for idx, col in enumerate(wrong_columns):
            Label(new_window, text=col).grid(row=idx + 1, column=0)
            if idx < len(missing_columns):
                entry = Entry(new_window)
                entry.grid(row=idx + 1, column=1)
                entry.insert(0, missing_columns[idx])
                entries.append(entry)
            else:
                Label(new_window, text="").grid(row=idx + 1, column=1)

        if len(missing_columns) > len(wrong_columns):
            Label(new_window, text="Add columns").grid(row=len(wrong_columns) + 1, column=0, columnspan=2)
            for idx in range(len(wrong_columns), len(missing_columns)):
                Label(new_window, text="").grid(row=idx + len(wrong_columns) + 2, column=0)
                entry = Entry(new_window)
                entry.insert(0, missing_columns[idx])
                entry.grid(row=idx + len(wrong_columns) + 2, column=1)
                entries.append(entry)

        Button(new_window, text="Correct column names", command=lambda: self.on_correct(wrong_columns, missing_columns,
                                                                                        table_name, entries,
                                                                                        new_window)).grid(
            row=len(wrong_columns) + len(missing_columns) + 2, columnspan=2)

        Button(new_window, text="Do not correct column names", command=lambda: self.do_not_correct(new_window)).grid(
            row=len(wrong_columns) + len(missing_columns) + 5, columnspan=2)

    def do_not_correct(self, new_window):
        new_window.destroy()

    def handle_validation(self, table_name):
        column_info = self.get_column_info(table_name)
        error_rows = []

        for index, row in self.df.iterrows():
            errors = self.validate_data(row, column_info, table_name)
            if errors:
                self.errors_found = True
                error_message = "; ".join([error[1] for error in errors])
                row_copy = row.copy()
                row_copy["Errors"] = error_message
                error_rows.append(row_copy)

        if not self.column_errors_found and error_rows:
            self.save_error_rows(self.df.columns, error_rows, table_name)
        else:
            self.remove_data_error_file(table_name)

        if not self.errors_found:
            messagebox.showinfo("Information", "No errors found, data correct!.")

    def check_data(self, excel_file, table_name):
        self.df = pd.read_excel(excel_file, dtype=object)
        self.df.replace(np.nan, None, regex=True, inplace=True)

        db_columns = self.get_columns(table_name)
        excel_columns = self.df.columns.tolist()

        wrong_columns = [col for col in excel_columns if col not in db_columns]
        missing_columns = [col for col in db_columns if col not in excel_columns]

        if wrong_columns or missing_columns:
            self.log_column_errors(wrong_columns, missing_columns, table_name)
            self.column_errors_found = True
            self.errors_found = True
            result = messagebox.askquestion("Question", "Do you want to correct column names?")
            if result == 'no':
                return
            else:
                self.correct_columns(wrong_columns, missing_columns, table_name)
        else:
            self.handle_validation(table_name)

    def validate_data(self, row, column_info, table_name):
        errors = []
        for column, value in row.items():
            column_details = column_info.get(column)
            if column_details is None:
                continue

            mysql_type = column_details["type"]
            is_nullable = column_details["null"] == "YES"
            is_unique = column_details.get("unique", False)
            python_type = type(value).__name__

            if not is_nullable and pd.isna(value):
                errors.append(("NOT NULL Column", f"{column}: Value cannot be null"))
                continue

            if value is None:
                continue

            if mysql_type.startswith("varchar"):
                max_length = int(mysql_type.split("(")[1].split(")")[0])
                if python_type != "str":
                    errors.append(("Invalid Data Type",
                                   f"{column}: Invalid data type (expected: varchar, received: {python_type})"))
                elif len(value) > max_length:
                    errors.append(("Exceeded Max Length",
                                   f"{column}: Exceeded max length (max: {max_length}, received: {len(value)})"))
            elif mysql_type.startswith("int") and python_type != "int":
                errors.append(
                    ("Invalid Data Type", f"{column}: Invalid data type (expected: int, received: {python_type})"))

            if is_unique and not pd.isna(value):
                if self.check_duplicate_in_file(column, value):
                    errors.append(("Duplicate value in the file", f"{column}: Duplicate value in the file '{value}'"))
                elif self.check_unique_constraint(table_name, column, value):
                    errors.append(("Duplicate value in the database", f"{column}: Duplicate value in the database '"
                                                                      f"{value}'"))

            if 'foreign_key' in column_details:
                referenced_table = column_details['foreign_key']['referenced_table']
                referenced_column = column_details['foreign_key']['referenced_column']
                if not self.check_foreign_key_constraint(referenced_table, referenced_column, value):
                    errors.append(("Foreign Key Violation",
                                   f"{column}: Value '{value}' does not exist in referenced table '"
                                   f"{referenced_table}.{referenced_column}'"))

        return errors

    def check_unique_constraint(self, table_name, column, value):
        cursor = self.connection.cursor()
        cursor.execute(f"SELECT COUNT(*) FROM {self.database}.{table_name} WHERE {column} = %s", (value,))
        count = cursor.fetchone()[0]
        cursor.close()
        return count > 0

    def check_duplicate_in_file(self, column, value):
        return self.df[self.df[column] == value].shape[0] > 1

    def check_foreign_key_constraint(self, referenced_table, referenced_column, value):
        cursor = self.connection.cursor()
        cursor.execute(f"SELECT COUNT(*) FROM {self.database}.{referenced_table} WHERE {referenced_column} ="
                       f" %s", (value,))
        count = cursor.fetchone()[0]
        cursor.close()
        return count > 0

    def save_error_rows(self, column_names, error_rows, table_name):
        log_folder = "errors"
        table_folder = os.path.join(log_folder, table_name)

        if not os.path.exists(table_folder):
            os.makedirs(table_folder)

        file_path = os.path.join(table_folder, "rows_validation_errors.xlsx")

        error_summary = {
            "Invalid Data Type": 0,
            "Exceeded Max Length": 0,
            "Duplicate value in the file": 0,
            "Duplicate value in the database": 0,
            "Value cannot be null": 0,
            "Unique Constraint Violation": 0,
            "Foreign Key Violation: value does not exist in referenced table": 0
        }

        for row in error_rows:
            errors = row["Errors"].split("; ")
            for error in errors:
                if "Invalid data type" in error:
                    error_summary["Invalid Data Type"] += 1
                elif "Exceeded max length" in error:
                    error_summary["Exceeded Max Length"] += 1
                elif "Duplicate value in the file" in error:
                    error_summary["Duplicate value in the file"] += 1
                elif "Duplicate value in the database" in error:
                    error_summary["Duplicate value in the database"] += 1
                elif "does not exist in referenced table" in error:
                    error_summary["Foreign Key Violation: value does not exist in referenced table"] += 1
                elif "Value cannot be null" in error:
                    error_summary["Value cannot be null"] += 1

        with pd.ExcelWriter(file_path) as writer:
            if error_rows:
                df_errors = pd.DataFrame(error_rows, columns=list(column_names) + ["Errors"])
                df_errors.insert(0, 'iRowId', range(1, len(df_errors) + 1))
                df_errors.to_excel(writer, sheet_name="Errors", index=False)

            summary_data = [{"Errors": key, "Quantity": value} for key, value in error_summary.items() if value > 0]
            if summary_data:
                all_errors_df = pd.DataFrame(summary_data)
                all_errors_df.to_excel(writer, sheet_name="ALL_ERRORS", index=False)

        messagebox.showinfo("Errors File Created",
                            f"Errors found. File created: {file_path}\nCheck the errors file for the details.")

    def log_column_errors(self, wrong_columns, missing_columns, table_name):
        script_dir = os.path.dirname(os.path.abspath(__file__))

        log_dir = os.path.join(script_dir, 'errors', table_name)
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)

        wb = Workbook()

        if wrong_columns:
            ws_wrong = wb.create_sheet("wrong column")
            ws_wrong.append(["Wrong Columns"])
            for col in wrong_columns:
                ws_wrong.append([col])

        if missing_columns:
            ws_missing = wb.create_sheet("missing columns")
            ws_missing.append(["Missing Columns"])
            for col in missing_columns:
                ws_missing.append([col])

        if 'Sheet' in wb.sheetnames and wb['Sheet'].max_row == 1:
            wb.remove(wb['Sheet'])

        if wb.sheetnames:
            log_file = os.path.join(log_dir, 'column_validation_errors.xlsx')
            wb.save(log_file)
            messagebox.showinfo("Errors File Created",
                                f"Errors found. File created: {log_file}\nCheck the errors file for details on the"
                                f" wrong and missing columns.")
        else:
            messagebox.showinfo("No Errors", "No errors were found in the columns.")

    def on_correct(self, wrong_columns, missing_columns, table_name, entries, new_window):
        for idx, col in enumerate(wrong_columns):
            if idx < len(missing_columns):
                new_name = entries[idx].get()
                if new_name:
                    self.df.rename(columns={col: new_name}, inplace=True)

        if len(missing_columns) > len(wrong_columns):
            for idx in range(len(wrong_columns), len(missing_columns)):
                self.df[missing_columns[idx]] = None

        db_columns = self.get_columns(table_name)
        self.df.drop(columns=[col for col in self.df.columns if col not in db_columns], inplace=True)

        self.column_errors_found = False
        self.errors_found = False

        new_window.destroy()
        messagebox.showinfo("Success", "Column names have been corrected")
        self.handle_validation(table_name)
        if self.on_import is not None:
            self.on_import(self.df, table_name)

    def remove_data_error_file(self, table_name):
        log_folder = "errors"
        table_folder = os.path.join(log_folder, table_name)
        log_file = os.path.join(table_folder, 'error_rows.xlsx')
        if os.path.exists(log_file):
            os.remove(log_file)
