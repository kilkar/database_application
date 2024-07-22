import shutil
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np


class ManageDataImport:
    def __init__(self, connection, database, data_store):
        self.connection = connection
        self.database = database
        self.tables = []
        self.data_store = data_store

    def import_data_df(self, df, table_name):
        try:
            cursor = self.connection.cursor()

            if not self.check_if_table_exists(table_name):
                self.data_store.create_key_for_db(table_name)

            for _, row in df.iterrows():
                placeholders = ", ".join(["%s"] * len(row))
                query = f"INSERT INTO {self.database}.{table_name} VALUES ({placeholders})"
                cursor.execute(query, tuple(row))
                self.append_row(table_name, row)

            self.save_to_history(table_name, df)

            self.connection.commit()
            cursor.close()
            messagebox.showinfo("Information", "The data has been imported into the database.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while importing data: {str(e)}")

    def check_if_table_exists(self, table_name):
        return self.data_store.check_if_table_exists(table_name)

    def append_row(self, table_name, row):
        self.data_store.append_row(table_name, row)

    def import_data(self, file_path, table_name):
        if not self.connection:
            messagebox.showerror("Error", "First connect to database.")
            return

        try:
            df = pd.read_excel(file_path, dtype=object)
            df.replace(np.nan, None, regex=True, inplace=True)
            cursor = self.connection.cursor()

            if not self.check_if_table_exists(table_name):
                self.data_store.create_key_for_db(table_name)

            for _, row in df.iterrows():
                placeholders = ", ".join(["%s"] * len(row))
                query = f"INSERT INTO {self.database}.{table_name} VALUES ({placeholders})"
                cursor.execute(query, tuple(row))
                self.append_row(table_name, row)

            self.save_to_history(table_name, df)

            self.connection.commit()
            cursor.close()
            messagebox.showinfo("Information", "The data has been imported into the database.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while importing data: {str(e)}")

    def save_to_history(self, table_name, df):
        history_folder = "history"
        if not os.path.exists(history_folder):
            os.makedirs(history_folder)

        excel_file_path = os.path.join(history_folder, f"{table_name}.xlsx")
        if os.path.exists(excel_file_path):
            wb = load_workbook(excel_file_path)
        else:
            wb = Workbook()
            del wb['Sheet']

        ws = wb.create_sheet(f"import_{datetime.now().strftime('%Y-%m-%d_%H-%M')}")

        ws.append(list(df.columns) + ['Import Date'])

        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r + [datetime.now().strftime('%Y-%m-%d_%H-%M')])

        wb.save(excel_file_path)

    def restore_initial_state(self):
        if not self.connection:
            messagebox.showerror("Error", "First connect to database.")
            return

        try:
            cursor = self.connection.cursor()

            for table_name, data_rows in reversed(self.data_store.get_items()):
                excel_file_path = os.path.join("history", f"{table_name}.xlsx")
                if os.path.exists(excel_file_path):
                    wb = load_workbook(excel_file_path)
                    ws = wb.create_sheet(f"deleted_{datetime.now().strftime('%Y-%m-%d_%H-%M')}")

                    ws.append(list(data_rows[0].index) + ['Deletion Date'])

                    for row in data_rows:
                        ws.append(row.tolist() + [datetime.now().strftime('%Y-%m-%d_%H-%M')])
                        wb.save(excel_file_path)

                        query_parts = []
                        values = []
                        for column, value in row.items():
                            if value is not None:
                                query_parts.append(f"{column} = %s")
                                values.append(value)

                        if query_parts:
                            query = f"DELETE FROM {self.database}.{table_name} WHERE {' AND '.join(query_parts)}"
                            cursor.execute(query, tuple(values))

            self.data_store.clear()
            self.connection.commit()
            cursor.close()
            messagebox.showinfo("Information", "The database has been restored to its initial state.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while restoring the database: {str(e)}")

    def save_history(self):
        history_folder = "history"
        if not os.path.exists(history_folder):
            messagebox.showerror("Error", "History folder does not exist!")
        else:
            new_folder = filedialog.asksaveasfilename(
                initialdir=os.getcwd(),
                title="Save as",
                initialfile="history",
                filetypes=[("Folder", "")],
            )
            if new_folder:
                try:
                    shutil.copytree(history_folder, new_folder)
                    messagebox.showinfo("Information", "The change history has been saved successfully.")
                except Exception as e:
                    messagebox.showerror("Error", f"An error occurred while saving the change history: {str(e)}")
